#!/usr/bin/env python3
"""Export per-SKU payloads from anunciar.xlsx using header-name mapping.

Creates one JSON file per SKU under `items_by_sku/` by default.
"""
import os
import re
import json
import argparse
import unicodedata
from decimal import Decimal

try:
    import openpyxl
    import pandas as pd
except Exception as e:
    raise SystemExit('pandas and openpyxl are required: pip install pandas openpyxl')


def normalize(s):
    if s is None:
        return ''
    s = str(s)
    s = unicodedata.normalize('NFKD', s)
    s = ''.join(c for c in s if not unicodedata.combining(c))
    return re.sub(r'\s+', ' ', s).strip()


def key_normal(s):
    return normalize(s).lower().replace(' ', '').replace('/', '').replace('-', '')


def load_config(path):
    with open(path, 'r', encoding='utf-8') as f:
        return json.load(f)


def find_sheet(workbook, config_name):
    # prefer a name that contains both 'livros' and 'fis'
    for name in workbook.sheetnames:
        low = name.lower()
        if 'livros' in low and 'fis' in low:
            return name
    # fallback exact
    if config_name in workbook.sheetnames:
        return config_name
    # last resort, try contains
    for name in workbook.sheetnames:
        if key_normal(config_name) in key_normal(name):
            return name
    return None


def find_anchor(sheet, anchor_texts):
    # look for any anchor text in row 1
    for cell in sheet[1]:
        if cell.value:
            v = normalize(cell.value).upper()
            for a in anchor_texts:
                if a.upper() in v:
                    return True
    return False


def build_header_maps_from_df(df_columns, mappings):
    # df_columns: list-like of header names
    norm_map = {key_normal(k): v for k, v in mappings.items()}
    header_to_field = {}
    for hdr in df_columns:
        if hdr is None:
            continue
        n = key_normal(hdr)
        if n in norm_map:
            header_to_field[hdr] = norm_map[n]
        else:
            for mk, field in norm_map.items():
                if mk in n or n in mk:
                    header_to_field[hdr] = field
                    break
    return header_to_field


def row_to_dict_from_series(s, header_names_map):
    # s is a pandas Series with original headers as keys
    return {hdr: s.get(hdr) for hdr in header_names_map.keys()}


def parse_prefix(val):
    if val is None:
        return None
    s = str(val).strip()
    parts = s.split(' - ', 1)
    return parts[0].strip()


def grams_to_kg(val):
    if val is None or val == '':
        return None
    s = str(val).strip().replace(',', '.')
    m = re.search(r'[\d\.]+', s)
    if not m:
        try:
            f = float(s)
        except Exception:
            return None
    else:
        f = float(m.group(0))
        # Treat input values as kilograms (no conversion); value already in kg
        return round(f, 3)
    return round(kg, 3)


def sanitize_filename(s):
    if s is None:
        return 'unknown'
    s = str(s)
    # keep alnum, dash, underscore
    return re.sub(r'[^A-Za-z0-9._-]', '_', s)


def digits_only(val):
    """Return only digit characters from val as a string, or None if empty/NA."""
    if val is None:
        return None
    try:
        if pd.isna(val):
            return None
    except Exception:
        pass
    # handle numeric types first to avoid losing precision when stringified
    if isinstance(val, float):
        # if float is integral, convert to int to avoid trailing .0
        if val.is_integer():
            s = str(int(val))
        else:
            s = ('%.15g' % val)
    else:
        s = str(val).strip()
    # remove non-digits
    digits = ''.join(re.findall(r"\d", s))
    return digits if digits else None


def main():
    p = argparse.ArgumentParser(description='Export per-SKU payloads from anunciar.xlsx')
    p.add_argument('--workbook', default='TESTE_ANUNCIOS/anunciar.xlsx')
    p.add_argument('--config', default='config/livros_info.json')
    p.add_argument('--dry-run', action='store_true')
    p.add_argument('--verbose', action='store_true')
    args = p.parse_args()

    cfg = load_config(args.config)

    # find sheet name first using openpyxl for robust matching
    wb = openpyxl.load_workbook(args.workbook, data_only=True)
    sheet_name = find_sheet(wb, cfg.get('sheet_name'))
    if sheet_name is None:
        raise SystemExit('Could not find a sheet matching "{}"'.format(cfg.get('sheet_name')))

    sheet = wb[sheet_name]
    if args.verbose:
        print('Using sheet:', sheet_name)

    # attempt to auto-detect the header row by scanning the first rows for mapping headers
    mappings = cfg.get('mappings', {}) or {}
    mapping_norms = [key_normal(k) for k in mappings.keys()]
    detected_header_row = None
    # scan first 60 rows for header-like matches
    for i, row in enumerate(sheet.iter_rows(min_row=1, max_row=60, values_only=True), start=1):
        found = 0
        for cell in row:
            if cell is None:
                continue
            n = key_normal(cell)
            for mk in mapping_norms:
                if mk in n or n in mk:
                    found += 1
                    break
        if args.verbose and found > 0:
            print(f'Row {i}: matched header-like cells =', found)
        # require at least 3 mapping matches to avoid false positives
        if found >= 3:
            detected_header_row = i
            break

    if detected_header_row:
        if args.verbose:
            print('Auto-detected header row at (1-based):', detected_header_row)
        header_row_idx = detected_header_row - 1
    else:
        header_row_idx = cfg.get('header_row', 2) - 1
        if args.verbose:
            print('Using configured header_row (1-based):', cfg.get('header_row', 2))

    # read with pandas using the detected header row
    df = pd.read_excel(args.workbook, sheet_name=sheet_name, header=header_row_idx, engine='openpyxl')

    # normalize df column names
    df.columns = [normalize(c) for c in df.columns]

    header_to_field = build_header_maps_from_df(df.columns, cfg.get('mappings', {}))
    if args.verbose:
        print('Mapped headers:')
        for h, f in header_to_field.items():
            print('  "{}" -> {}'.format(h, f))

    # prepare field->header
    field_to_header = {v: k for k, v in header_to_field.items()}

    grouped = {}

    for _, row in df.iterrows():
        # find SKU
        sku_header = field_to_header.get('sku')
        if not sku_header:
            continue
        sku = row.get(sku_header)
        if pd.isna(sku):
            continue

        title = row.get(field_to_header.get('title')) if field_to_header.get('title') else None
        ncm = row.get(field_to_header.get('ncm')) if field_to_header.get('ncm') else None
        ean = row.get(field_to_header.get('ean')) if field_to_header.get('ean') else None
        origin_val = row.get(field_to_header.get('origin_detail')) if field_to_header.get('origin_detail') else None
        csosn_val = row.get(field_to_header.get('csosn')) if field_to_header.get('csosn') else None

        weight_header = field_to_header.get('net_weight')
        weight_val = row.get(weight_header) if weight_header else None
        if weight_val is None or (isinstance(weight_val, float) and pd.isna(weight_val)):
            # heuristic
            for h in df.columns:
                if 'peso' in h.lower():
                    weight_val = row.get(h)
                    break

        net_weight = grams_to_kg(weight_val)

        origin_prefix = parse_prefix(origin_val)
        csosn_prefix = parse_prefix(csosn_val)

        payload = {
            'sku': str(sku).strip(),
            'title': normalize(title),
            'type': cfg.get('constants', {}).get('type', 'single'),
            'measurement_unit': cfg.get('constants', {}).get('measurement_unit', 'UN'),
            'tax_information': {
                'ncm': digits_only(ncm),
                'origin_type': cfg.get('constants', {}).get('origin_type', 'reseller'),
                'origin_detail': origin_prefix if origin_prefix is not None else None,
                'tax_rule_id': None,
                'csosn': csosn_prefix if csosn_prefix is not None else None,
                'ean': digits_only(ean),
                'net_weight': net_weight
            }
        }

        grouped.setdefault(str(sku).strip(), []).append(payload)

    out_dir = cfg.get('output', {}).get('dir', 'items_by_sku')
    filename_template = cfg.get('output', {}).get('filename_template', '{sku}.json')

    if args.dry_run:
        print('Dry run: found {} SKUs'.format(len(grouped)))
        for i, (sku, items) in enumerate(grouped.items()):
            print('\nSKU:', sku, 'rows:', len(items))
            print(json.dumps(items[0], ensure_ascii=False, indent=2))
            if i >= 4:
                break
        return

    os.makedirs(out_dir, exist_ok=True)
    written = 0
    for sku, items in grouped.items():
        safe = sanitize_filename(sku)
        path = os.path.join(out_dir, filename_template.format(sku=safe))
        data = items[0] if len(items) == 1 else items
        with open(path, 'w', encoding='utf-8') as f:
            json.dump(data, f, ensure_ascii=False, indent=2)
        written += 1

    print('Wrote {} files to {}'.format(written, out_dir))


if __name__ == '__main__':
    main()
